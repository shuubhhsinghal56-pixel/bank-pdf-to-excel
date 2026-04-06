from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import io
import re
import json
import logging
import uuid
import traceback
from pathlib import Path
from typing import List, Optional, Tuple
from pydantic import BaseModel
from datetime import datetime, timezone

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
except Exception:
    LlmChat = None
    UserMessage = None

# Configure litellm to have shorter timeouts and no retries (we handle retries ourselves)
try:
    import litellm
    litellm.request_timeout = 30  # 30 second timeout per request
    litellm.num_retries = 0  # No internal retries
    litellm.num_retries_per_request = 0
    # Also set at openai client level
    import openai
    openai.timeout = 30
except Exception:
    pass

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')
EMERGENT_MODEL_PROVIDER = os.environ.get('EMERGENT_MODEL_PROVIDER', 'anthropic')
EMERGENT_MODEL_NAME = os.environ.get('EMERGENT_MODEL_NAME', 'claude-3-5-haiku-20241022')

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 25 * 1024 * 1024  # 25MB
BALANCE_TOLERANCE = 0.051
DATE_TOKEN_RE = re.compile(r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})\b')
LINE_START_DATE_RE = re.compile(r'^\s*(?:\d+\s+)?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})\b')
AMOUNT_RE = re.compile(
    r'(?<!\S)(?:₹\s*)?\(?-?\d[\d,]*(?:\.\d+)?\)?(?:\s*(?:CR|DR|\(CR\)|\(DR\)))?(?!\S)',
    re.IGNORECASE,
)
DECIMAL_AMOUNT_RE = re.compile(
    r'(?<!\S)(?:₹\s*)?\(?-?\d[\d,]*\.\d+\)?(?:\s*(?:CR|DR|\(CR\)|\(DR\)))?(?!\S)',
    re.IGNORECASE,
)


# ─── Models ───────────────────────────────────────────────────────────────────

class DetectedFormat(BaseModel):
    bank_name: str = ""
    columns: List[str] = []
    date_format: str = ""
    amount_style: str = ""  # "separate" or "combined"
    currency_symbol: str = "₹"
    raw_text_preview: str = ""

class Transaction(BaseModel):
    date: str = ""
    narration: str = ""
    reference: str = ""
    debit: Optional[float] = None
    credit: Optional[float] = None
    balance: Optional[float] = None
    balance_mismatch: bool = False
    page_number: int = 0

class ParseResult(BaseModel):
    transactions: List[Transaction] = []
    summary: dict = {}
    errors: List[str] = []
    total_pages: int = 0
    pages_processed: int = 0


class ParseConfidence(BaseModel):
    score: int = 0
    level: str = "low"
    review_recommended: bool = True
    reasons: List[str] = []
    pages_with_transactions: int = 0
    skipped_pages: int = 0
    locally_parsed_pages: int = 0
    ai_fallback_pages: int = 0
    token_strategy: str = "local_only"

class FormatConfirmation(BaseModel):
    bank_name: str
    columns: List[str]
    date_format: str
    amount_style: str
    currency_symbol: str
    full_text: str  # base64 encoded full text
    total_pages: int

class ExcelRequest(BaseModel):
    transactions: List[dict]
    summary: dict


# ─── Helpers ──────────────────────────────────────────────────────────────────

def parse_indian_number(s: str) -> Optional[float]:
    """Parse Indian number format like 1,00,000.00 or plain numbers."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    if not s or s == '-' or s.lower() == 'null' or s.lower() == 'none' or s.lower() == 'n/a':
        return None
    # Remove currency symbols
    s = re.sub(r'[₹$€£]', '', s).strip()
    # Remove Dr/Cr suffix
    dr_cr = None
    if s.upper().endswith('(DR)') or s.upper().endswith('DR'):
        dr_cr = 'dr'
        s = re.sub(r'\s*\(?\s*DR\s*\)?\s*$', '', s, flags=re.IGNORECASE).strip()
    elif s.upper().endswith('(CR)') or s.upper().endswith('CR'):
        dr_cr = 'cr'
        s = re.sub(r'\s*\(?\s*CR\s*\)?\s*$', '', s, flags=re.IGNORECASE).strip()
    # Remove commas
    s = s.replace(',', '')
    # Handle negative numbers
    is_negative = False
    if s.startswith('-') or s.startswith('('):
        is_negative = True
        s = s.strip('-').strip('(').strip(')')
    try:
        val = float(s)
        if is_negative:
            val = -val
        return val
    except ValueError:
        return None


def format_indian_number(num: float) -> str:
    """Format number in Indian style: ₹1,00,000.00"""
    if num is None:
        return ""
    is_negative = num < 0
    num = abs(num)
    integer_part = int(num)
    decimal_part = round(num - integer_part, 2)
    decimal_str = f"{decimal_part:.2f}"[1:]  # .00
    s = str(integer_part)
    if len(s) <= 3:
        result = s
    else:
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + ',' + result
            s = s[:-2]
    sign = '-' if is_negative else ''
    return f"{sign}₹{result}{decimal_str}"


def extract_json_from_response(text: str) -> dict:
    """Extract JSON from Claude response that may contain markdown code blocks."""
    # Try direct parse first
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    # Try extracting from code blocks
    patterns = [
        r'```json\s*([\s\S]*?)\s*```',
        r'```\s*([\s\S]*?)\s*```',
        r'\{[\s\S]*\}',
        r'\[[\s\S]*\]',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            try:
                return json.loads(match.group(1) if '```' in pattern else match.group(0))
            except (json.JSONDecodeError, IndexError):
                continue
    raise ValueError(f"Could not extract JSON from response: {text[:500]}")


def is_ai_available() -> bool:
    return bool(EMERGENT_LLM_KEY and LlmChat and UserMessage)


async def call_claude(prompt: str, system_msg: str = "", timeout_seconds: int = 60, max_retries: int = 2) -> str:
    """Call Claude via emergentintegrations with timeout and retry."""
    import asyncio

    if not is_ai_available():
        raise RuntimeError("AI parser is unavailable because EMERGENT_LLM_KEY is missing or the SDK is not installed.")
    
    last_error = None
    for attempt in range(max_retries):
        try:
            session_id = str(uuid.uuid4())
            chat = LlmChat(
                api_key=EMERGENT_LLM_KEY,
                session_id=session_id,
                system_message=system_msg or "You are a bank statement parsing expert. Always respond with valid JSON only, no markdown, no explanation."
            )
            chat.with_model(EMERGENT_MODEL_PROVIDER, EMERGENT_MODEL_NAME)
            user_msg = UserMessage(text=prompt)
            response = await asyncio.wait_for(
                chat.send_message(user_msg),
                timeout=timeout_seconds
            )
            return response
        except asyncio.TimeoutError:
            last_error = TimeoutError(f"AI call timed out after {timeout_seconds}s (attempt {attempt+1}/{max_retries})")
            logger.warning(f"AI call timeout attempt {attempt+1}/{max_retries}")
            continue
        except Exception as e:
            error_str = str(e)
            if "budget" in error_str.lower() or "exceeded" in error_str.lower():
                raise BudgetExceededError("AI budget exceeded. Please top up your Universal Key balance at Profile → Universal Key → Add Balance.")
            last_error = e
            logger.warning(f"AI call error attempt {attempt+1}/{max_retries}: {error_str[:200]}")
            if attempt < max_retries - 1:
                await asyncio.sleep(2)  # Short wait before retry
                continue
            raise
    
    raise last_error or Exception("AI call failed after all retries")


class BudgetExceededError(Exception):
    pass


def normalize_whitespace(text: str) -> str:
    return re.sub(r'\s+', ' ', text or '').strip()


def normalize_date(date_str: str) -> str:
    raw = (date_str or "").strip()
    if not raw:
        return ""

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%d %m %Y", "%d %b %Y", "%d %B %Y"):
        try:
            normalized_raw = raw if "%b" in fmt or "%B" in fmt else raw.replace("-", "/")
            normalized_fmt = fmt if "%b" in fmt or "%B" in fmt else fmt.replace("-", "/")
            return datetime.strptime(normalized_raw, normalized_fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return raw.replace("/", "-")


def detect_bank_name_locally(text: str) -> str:
    upper_text = text.upper()
    if (
        "HDFCBANKLTD" in upper_text
        or "HDFCBANKLIMITED" in upper_text
        or "HDFC BANK LIMITED" in upper_text
        or "HDFC000" in upper_text
    ):
        return "HDFC Bank"
    if "BANK OF MAHARASHTRA" in upper_text:
        return "Bank Of Maharashtra"
    if re.search(r'\bUBIN\d{4,}\b', text, flags=re.IGNORECASE):
        return "Union Bank"
    if re.search(r'\bKKBK\d{4,}\b', text, flags=re.IGNORECASE):
        return "Kotak Bank"

    statement_match = re.search(r'Statement of\s+([A-Za-z ]+?)\s+Account', text, flags=re.IGNORECASE)
    if statement_match:
        bank_name = normalize_whitespace(statement_match.group(1)).title()
        if bank_name.lower() not in {"account", "statement", "sb"}:
            if "Bank" not in bank_name:
                bank_name = f"{bank_name} Bank"
            return bank_name

    for line in text.splitlines():
        cleaned = normalize_whitespace(line)
        if not cleaned:
            continue
        upper = cleaned.upper()
        if "STATEMENT" in upper or "ACCOUNT NO" in upper or "PERIOD" in upper:
            continue
        if "BANK" in upper and len(cleaned) <= 80:
            return cleaned.title()
    return "Unknown Bank"


def detect_columns_locally(text: str) -> List[str]:
    candidate_headers = [
        ["Date", "Narration", "Reference", "Debit", "Credit", "Balance"],
        ["Date", "Description", "Reference", "Debit", "Credit", "Balance"],
        ["Date", "Transaction Id", "Particulars", "Amount", "Balance"],
    ]

    for line in text.splitlines():
        cleaned = normalize_whitespace(line).lower()
        if "date" in cleaned and ("balance" in cleaned or "deposit" in cleaned or "withdrawal" in cleaned):
            headers = []
            if "date" in cleaned:
                headers.append("Date")
            if "transaction id" in cleaned:
                headers.append("Transaction Id")
            if "narration" in cleaned or "particular" in cleaned or "description" in cleaned:
                headers.append("Narration")
            if "ref" in cleaned or "chq" in cleaned or "cheque" in cleaned:
                headers.append("Reference")
            if "amount(" in cleaned or " amount " in f" {cleaned} ":
                headers.append("Amount")
            if "withdrawal" in cleaned or "debit" in cleaned:
                headers.append("Debit")
            if "deposit" in cleaned or "credit" in cleaned:
                headers.append("Credit")
            if "balance" in cleaned:
                headers.append("Balance")
            if len(headers) >= 4:
                return headers

    return candidate_headers[0]


def detect_amount_style_locally(text: str) -> str:
    upper = text.upper()
    if "WITHDRAWAL" in upper and "DEPOSIT" in upper:
        return "separate"
    if re.search(r'\d[\d,]*(?:\.\d+)?\s*(?:CR|DR|\(CR\)|\(DR\))', upper):
        return "combined"
    if "AMOUNT(" in upper and "BALANCE(" in upper:
        return "combined"
    if "WITHDRAWAL" in upper or "DEBIT" in upper or "DEPOSIT" in upper or "CREDIT" in upper:
        return "separate"
    return "separate"


def detect_date_format_locally(text: str) -> str:
    match = DATE_TOKEN_RE.search(text)
    if not match:
        return "DD/MM/YYYY"
    token = match.group(0)
    if "-" in token:
        return "DD-MM-YYYY" if len(token.split("-")[-1]) == 4 else "DD-MM-YY"
    return "DD/MM/YYYY" if len(token.split("/")[-1]) == 4 else "DD/MM/YY"


def detect_format_locally(text: str) -> dict:
    return {
        "bank_name": detect_bank_name_locally(text),
        "columns": detect_columns_locally(text),
        "date_format": detect_date_format_locally(text),
        "amount_style": detect_amount_style_locally(text),
        "currency_symbol": "₹" if "₹" in text else "₹",
    }


def split_text_into_pages(full_text: str) -> List[Tuple[int, str]]:
    pages: List[Tuple[int, str]] = []
    current_page_number: Optional[int] = None
    current_lines: List[str] = []

    for raw_line in full_text.splitlines():
        page_match = re.match(r'^--- PAGE (\d+) ---$', raw_line.strip())
        if page_match:
            if current_page_number is not None:
                pages.append((current_page_number, "\n".join(current_lines).strip()))
            current_page_number = int(page_match.group(1))
            current_lines = []
            continue
        if current_page_number is None:
            current_page_number = 1
        current_lines.append(raw_line)

    if current_page_number is not None:
        pages.append((current_page_number, "\n".join(current_lines).strip()))

    return [(page_no, text) for page_no, text in pages if text]


def sanitize_page_text(page_text: str) -> str:
    lines = [line.rstrip() for line in page_text.splitlines()]
    if not lines:
        return page_text

    first_date_index: Optional[int] = None
    for idx, line in enumerate(lines):
        if LINE_START_DATE_RE.match(normalize_whitespace(line)):
            first_date_index = idx
            break

    header_index: Optional[int] = None
    for idx, line in enumerate(lines[:30]):
        upper = normalize_whitespace(line).upper()
        if (
            "STATEMENTOF ACCOUNT" in upper
            or "STATEMENT OF ACCOUNT" in upper
            or "ACCOUNTBRANCH :" in upper
            or "ACCOUNTBRANCH:" in upper
        ):
            header_index = idx

    if header_index is not None and first_date_index is not None and first_date_index > header_index:
        lines = lines[first_date_index:]

    footer_markers = (
        "*CLOSINGBALANCEINCLUDESFUNDS",
        "*CLOSING BALANCE INCLUDES FUNDS",
        "CONTENTSOFTHISSTATEMENTWILLBECONSIDEREDCORRECT",
        "CONTENTS OF THIS STATEMENT WILL BE CONSIDERED CORRECT",
        "STATEACCOUNTBRANCHGSTN",
        "REGISTEREDOFFICEADDRESS:HDFCBANKHOUSE",
        "REGISTERED OFFICE ADDRESS:HDFC BANK HOUSE",
    )
    trimmed_lines: List[str] = []
    for line in lines:
        upper = normalize_whitespace(line).upper()
        if any(marker in upper for marker in footer_markers):
            break
        if upper == "HDFCBANKLIMITED" and trimmed_lines:
            break
        trimmed_lines.append(line)

    return "\n".join(trimmed_lines).strip()


def page_has_transaction_markers(page_text: str) -> bool:
    lines = [normalize_whitespace(line) for line in page_text.splitlines() if normalize_whitespace(line)]
    transaction_like_lines = 0

    for line in lines:
        if is_non_transaction_line(line):
            continue
        if "OPENING BALANCE" in line.upper() or "CLOSING BALANCE" in line.upper():
            continue

        amount_match_count = len(DECIMAL_AMOUNT_RE.findall(line))
        if LINE_START_DATE_RE.match(line) and amount_match_count >= 1:
            transaction_like_lines += 1
            continue

        if re.match(r'^\d+\s+\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\b', line) and amount_match_count >= 1:
            transaction_like_lines += 1
            continue

    return transaction_like_lines >= 1


def is_non_transaction_line(line: str) -> bool:
    upper = normalize_whitespace(line).upper()
    if not upper:
        return True

    ignored_prefixes = (
        "ACCOUNT STATEMENT",
        "SAVINGS ACCOUNT TRANSACTIONS",
        "CURRENT ACCOUNT",
        "YOUR DETAILS",
        "ACCOUNT DETAILS",
        "STATEMENT DETAILS",
        "ACCOUNT SUMMARY",
        "PAGENO.:",
        "ACCOUNTBRANCH :",
        "ACCOUNTBRANCH:",
        "ADDRESS :",
        "CITY :",
        "STATE :",
        "FLATNO",
        "MYWOODS",
        "CURRENCY :",
        "EMAIL :",
        "A/COPENDATE :",
        "JOINTHOLDERS:",
        "RTGS/NEFTIFSC:",
        "BRANCHCODE :",
        "NOMINATION:",
        "FROM :",
        "DATE NARRATION CHQ./REF.NO.",
        "DATE NARRATION CHQ/REF NO",
        "DATE NARRATION CHQ NO",
        "IMPORTANT INFORMATION",
        "STATEMENT GENERATED ON",
        "# DATE DESCRIPTION",
        "TRAN DATE",
        "DATE TRANSACTION ID",
        "OPENING BALANCE",
        "TRANSACTION TOTAL",
        "CLOSING BALANCE",
        "UNLESS THE CONSTITUENT",
        "THE CLOSING BALANCE AS",
        "WE WOULD LIKE TO REITERATE",
        "DEPOSIT INSURANCE",
        "IN COMPLIANCE WITH",
        "TO ENSURE YOU NEVER",
        "THIS IS A SYSTEM GENERATED",
        "++++ END OF STATEMENT",
        "ICONN-",
        "VMT-",
        "AUTOSWEEP-",
        "REV SWEEP-",
        "SWEEP TRF-",
        "CWDR-",
        "PUR-",
        "TIP/",
        "RATE.DIFF-",
        "CLG-",
        "EDC-",
        "SETU ",
        "INT.PD-",
        "INT.COLL-",
    )
    return upper.startswith(ignored_prefixes)


def extract_opening_balance(line: str) -> Optional[float]:
    upper = normalize_whitespace(line).upper()
    if "OPENING BALANCE" not in upper or LINE_START_DATE_RE.match(line):
        return None
    match = re.search(r'((?:₹\s*)?\d[\d,]*\.\d{2})$', line, flags=re.IGNORECASE)
    if not match:
        return None
    return parse_indian_number(match.group(1))


def extract_particulars_and_amounts(text: str, amount_style: str) -> Tuple[str, List[str]]:
    raw = normalize_whitespace(text)
    if not raw:
        return "", []

    matcher = AMOUNT_RE if amount_style == "combined" else DECIMAL_AMOUNT_RE
    amount_matches = list(matcher.finditer(raw))
    if not amount_matches:
        return raw, []

    particulars = raw[:amount_matches[0].start()].strip()
    amounts = [match.group(0) for match in amount_matches]
    return particulars, amounts


def evaluate_transaction_quality(transactions: List[Transaction]) -> Tuple[int, int]:
    if not transactions:
        return 0, 0

    mismatches = 0
    usable_rows = 0
    previous_balance: Optional[float] = None

    for txn in transactions:
        if txn.balance is not None:
            usable_rows += 1
        if previous_balance is not None and txn.balance is not None:
            expected = previous_balance
            if txn.debit and txn.debit > 0:
                expected -= txn.debit
            if txn.credit and txn.credit > 0:
                expected += txn.credit
            if abs(expected - txn.balance) > BALANCE_TOLERANCE:
                mismatches += 1
        if txn.balance is not None:
            previous_balance = txn.balance

    return usable_rows, mismatches


def calculate_parse_score(transactions: List[Transaction]) -> int:
    usable_rows, mismatches = evaluate_transaction_quality(transactions)
    long_row_penalty = sum(1 for txn in transactions if len(normalize_whitespace(txn.narration)) > 90)
    empty_particular_penalty = sum(1 for txn in transactions if not normalize_whitespace(txn.narration))
    return (usable_rows * 100) - (mismatches * 25) - (long_row_penalty * 20) - (empty_particular_penalty * 50)


def build_parse_confidence(
    transactions: List[Transaction],
    total_pages: int,
    pages_with_transactions: int,
    skipped_pages: int,
    locally_parsed_pages: int,
    ai_fallback_pages: int,
    errors: List[str],
) -> ParseConfidence:
    if not transactions:
        return ParseConfidence(
            score=0,
            level="low",
            review_recommended=True,
            reasons=["No transactions could be parsed from the extracted PDF text."],
            pages_with_transactions=pages_with_transactions,
            skipped_pages=skipped_pages,
            locally_parsed_pages=locally_parsed_pages,
            ai_fallback_pages=ai_fallback_pages,
            token_strategy="ai_fallback_only" if ai_fallback_pages else "local_only",
        )

    usable_rows, mismatches = evaluate_transaction_quality(transactions)
    mismatch_rate = (mismatches / usable_rows) if usable_rows else 1.0
    page_coverage = (pages_with_transactions / total_pages) if total_pages else 0.0
    ai_ratio = (ai_fallback_pages / pages_with_transactions) if pages_with_transactions else 0.0

    score = 100
    score -= min(55, round(mismatch_rate * 100))
    score -= min(20, max(0, len(errors) - ai_fallback_pages) * 4)
    score -= min(18, ai_fallback_pages * 3)
    score -= min(10, max(0, pages_with_transactions - locally_parsed_pages - ai_fallback_pages) * 2)
    if page_coverage < 0.5:
        score -= 20
    elif page_coverage < 0.8:
        score -= 8
    score = max(0, min(100, score))

    reasons: List[str] = []
    if mismatches == 0:
        reasons.append("Running balances reconciled cleanly across parsed rows.")
    else:
        reasons.append(f"{mismatches} row(s) do not reconcile with the running balance.")

    if ai_fallback_pages == 0:
        reasons.append("Parsing completed locally, which keeps token usage minimal.")
    else:
        reasons.append(
            f"AI fallback was used on {ai_fallback_pages} transaction page(s) only after local parsing failed."
        )

    if skipped_pages > 0:
        reasons.append(f"{skipped_pages} non-transaction page(s) were skipped automatically.")

    if score >= 85 and mismatches == 0:
        level = "high"
    elif score >= 65:
        level = "medium"
    else:
        level = "low"

    return ParseConfidence(
        score=score,
        level=level,
        review_recommended=(level == "low" or mismatches > 0),
        reasons=reasons,
        pages_with_transactions=pages_with_transactions,
        skipped_pages=skipped_pages,
        locally_parsed_pages=locally_parsed_pages,
        ai_fallback_pages=ai_fallback_pages,
        token_strategy="ai_fallback_only" if ai_fallback_pages else "local_only",
    )


def split_amounts(
    amount_matches: List[str],
    amount_style: str,
    previous_balance: Optional[float] = None,
) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if not amount_matches:
        return None, None, None

    cleaned = [normalize_whitespace(match) for match in amount_matches]
    balance = parse_indian_number(cleaned[-1])
    debit = None
    credit = None
    movement_tokens = cleaned[:-1]

    if amount_style == "combined" and movement_tokens:
        token = movement_tokens[-1]
        value = abs(parse_indian_number(token) or 0)
        upper = token.upper()
        if "CR" in upper:
            credit = value
        elif "DR" in upper:
            debit = value
        elif parse_indian_number(token) and parse_indian_number(token) < 0:
            debit = value
        else:
            debit = value
        return debit, credit, balance

    if len(movement_tokens) >= 2:
        debit = parse_indian_number(movement_tokens[0])
        credit = parse_indian_number(movement_tokens[1])
    elif len(movement_tokens) == 1:
        token = movement_tokens[0]
        value = parse_indian_number(token)
        upper = token.upper()
        if value is not None:
            amount = abs(value)
            if previous_balance is not None and balance is not None:
                if abs((previous_balance - amount) - balance) <= BALANCE_TOLERANCE:
                    debit = amount
                    credit = None
                elif abs((previous_balance + amount) - balance) <= BALANCE_TOLERANCE:
                    credit = amount
                    debit = None
                elif "CR" in upper:
                    credit = amount
                elif "DR" in upper or value < 0:
                    debit = amount
                else:
                    debit = amount
            elif "CR" in upper:
                credit = abs(value)
            elif "DR" in upper or value < 0:
                debit = abs(value)
            else:
                debit = abs(value)

    if debit is not None:
        debit = abs(debit)
    if credit is not None:
        credit = abs(credit)
    return debit, credit, balance


def build_transaction_from_parts(
    date_text: str,
    particulars: str,
    amount_matches: List[str],
    amount_style: str,
    page_number: int,
    previous_balance: Optional[float] = None,
) -> Optional[Transaction]:
    particulars = DATE_TOKEN_RE.sub("", normalize_whitespace(particulars)).strip()
    debit, credit, balance = split_amounts(
        amount_matches,
        amount_style,
        previous_balance=previous_balance,
    )
    if balance is None:
        return None

    return Transaction(
        date=normalize_date(date_text),
        narration=particulars,
        reference="",
        debit=debit,
        credit=credit,
        balance=balance,
        page_number=page_number,
    )


def parse_transaction_line(
    line: str,
    amount_style: str,
    previous_balance: Optional[float] = None,
) -> Optional[Transaction]:
    raw = normalize_whitespace(line)
    date_match = LINE_START_DATE_RE.match(raw)
    if not raw or not date_match:
        return None

    date = normalize_date(date_match.group(1))
    remainder = raw[date_match.end():].strip()
    particulars, amounts = extract_particulars_and_amounts(remainder, amount_style)
    if not amounts:
        return None
    particulars = DATE_TOKEN_RE.sub("", particulars).strip()

    debit, credit, balance = split_amounts(amounts, amount_style, previous_balance=previous_balance)
    if balance is None:
        return None

    return Transaction(
        date=date,
        narration=particulars,
        reference="",
        debit=debit,
        credit=credit,
        balance=balance,
        page_number=0,
    )


def parse_transaction_block(
    block_lines: List[str],
    amount_style: str,
    page_number: int,
    previous_balance: Optional[float] = None,
) -> Optional[Transaction]:
    if not block_lines:
        return None

    date_line_index = None
    date_match = None
    for idx, line in enumerate(block_lines):
        match = LINE_START_DATE_RE.match(line)
        if match:
            date_line_index = idx
            date_match = match
            break

    if date_line_index is None or date_match is None:
        return None

    date_text = date_match.group(1)
    prefix_lines = [line for line in block_lines[:date_line_index] if not is_non_transaction_line(line)]
    date_line = block_lines[date_line_index]
    suffix_lines = [line for line in block_lines[date_line_index + 1:] if not is_non_transaction_line(line)]
    date_remainder = date_line[date_match.end():].strip()

    particulars_on_line, amount_matches = extract_particulars_and_amounts(date_remainder, amount_style)
    if not amount_matches:
        combined_text = normalize_whitespace(" ".join([date_remainder, *suffix_lines]))
        particulars_on_line, amount_matches = extract_particulars_and_amounts(combined_text, amount_style)

    if not amount_matches:
        return None

    particulars_parts = [*prefix_lines]
    if particulars_on_line:
        particulars_parts.append(particulars_on_line)
    particulars_parts.extend(suffix_lines)
    particulars = normalize_whitespace(" ".join(particulars_parts))

    return build_transaction_from_parts(
        date_text=date_text,
        particulars=particulars,
        amount_matches=amount_matches,
        amount_style=amount_style,
        page_number=page_number,
        previous_balance=previous_balance,
    )


def parse_transactions_blockwise(full_text: str, amount_style: str) -> List[Transaction]:
    transactions: List[Transaction] = []
    current_page = 1
    pending_prefix_lines: List[str] = []
    current_block: List[str] = []
    previous_balance: Optional[float] = None

    def finalize_block() -> None:
        nonlocal current_block, previous_balance
        if not current_block:
            return
        txn = parse_transaction_block(
            current_block,
            amount_style,
            page_number=current_page,
            previous_balance=previous_balance,
        )
        current_block = []
        if txn:
            transactions.append(txn)
            if txn.balance is not None:
                previous_balance = txn.balance

    for raw_line in full_text.splitlines():
        line = normalize_whitespace(raw_line)
        if raw_line.startswith("--- PAGE"):
            finalize_block()
            page_match = re.search(r'PAGE\s+(\d+)', raw_line)
            if page_match:
                current_page = int(page_match.group(1))
            pending_prefix_lines = []
            continue

        if not line or is_non_transaction_line(line):
            if line and (
                "TRAN DATE" in line.upper()
                or "DATE TRANSACTION ID" in line.upper()
                or "STATEMENT OF" in line.upper()
            ):
                pending_prefix_lines = []
            opening_match = re.search(r'OPENING BALANCE.*?((?:₹\s*)?\d[\d,]*\.\d{2})$', line, flags=re.IGNORECASE)
            if opening_match:
                finalize_block()
                pending_prefix_lines = []
                previous_balance = parse_indian_number(opening_match.group(1))
            continue

        opening_balance = extract_opening_balance(line)
        if opening_balance is not None:
            finalize_block()
            pending_prefix_lines = []
            previous_balance = opening_balance
            continue

        if LINE_START_DATE_RE.match(line):
            finalize_block()
            current_block = [*pending_prefix_lines, line]
            pending_prefix_lines = []
            continue

        if current_block:
            current_block.append(line)
        else:
            pending_prefix_lines.append(line)

    finalize_block()
    return transactions


def parse_transactions_linewise(full_text: str, amount_style: str) -> List[Transaction]:
    transactions: List[Transaction] = []
    current_page = 1
    pending_prefix_lines: List[str] = []
    previous_balance: Optional[float] = None
    first_transaction_seen_on_page = False

    for raw_line in full_text.splitlines():
        line = normalize_whitespace(raw_line)
        if raw_line.startswith("--- PAGE"):
            page_match = re.search(r'PAGE\s+(\d+)', raw_line)
            if page_match:
                current_page = int(page_match.group(1))
            pending_prefix_lines = []
            first_transaction_seen_on_page = False
            continue

        if not line or is_non_transaction_line(line):
            if line and (
                "TRAN DATE" in line.upper()
                or "DATE TRANSACTION ID" in line.upper()
                or "STATEMENT OF" in line.upper()
            ):
                pending_prefix_lines = []
            opening_match = re.search(r'OPENING BALANCE.*?((?:₹\s*)?\d[\d,]*\.\d{2})$', line, flags=re.IGNORECASE)
            if opening_match:
                pending_prefix_lines = []
                previous_balance = parse_indian_number(opening_match.group(1))
            continue

        opening_balance = extract_opening_balance(line)
        if opening_balance is not None:
            pending_prefix_lines = []
            previous_balance = opening_balance
            continue

        if LINE_START_DATE_RE.match(line):
            date_match = LINE_START_DATE_RE.match(line)
            if not date_match:
                continue

            remainder = line[date_match.end():].strip()
            line_particulars, line_amounts = extract_particulars_and_amounts(remainder, amount_style)
            particulars = normalize_whitespace(" ".join([*pending_prefix_lines, line_particulars]))
            pending_prefix_lines = []

            if not line_amounts:
                continue

            txn = build_transaction_from_parts(
                date_text=date_match.group(1),
                particulars=particulars,
                amount_matches=line_amounts,
                amount_style=amount_style,
                page_number=current_page,
                previous_balance=previous_balance,
            )
            if not txn:
                continue
            transactions.append(txn)
            previous_balance = txn.balance
            first_transaction_seen_on_page = True
            continue

        if first_transaction_seen_on_page and transactions and transactions[-1].page_number == current_page:
            transactions[-1].narration = normalize_whitespace(f"{transactions[-1].narration} {line}")
        else:
            pending_prefix_lines.append(line)

    return transactions


def parse_transactions_locally(full_text: str, amount_style: str) -> List[Transaction]:
    strategies = [
        parse_transactions_linewise(full_text, amount_style),
        parse_transactions_blockwise(full_text, amount_style),
    ]
    strategies = [strategy for strategy in strategies if strategy]
    if not strategies:
        return []
    return max(strategies, key=calculate_parse_score)


async def parse_page_with_ai(
    page_text: str,
    page_number: int,
    data: FormatConfirmation,
) -> List[Transaction]:
    columns_desc = ', '.join(data.columns)
    amount_desc = "combined in one column with Dr/Cr suffix" if data.amount_style == "combined" else "in separate Debit and Credit columns"
    parse_prompt = f"""Parse the following bank statement page into transactions.

Bank: {data.bank_name}
Column order: {columns_desc}
Date format: {data.date_format}
Amount style: Amounts are {amount_desc}
Currency: {data.currency_symbol}
Page number: {page_number}

IMPORTANT RULES:
1. Extract ONLY actual transaction rows. Skip headers, footers, page numbers, summary rows, and opening/closing balance lines.
2. Normalize all dates to DD-MM-YYYY format.
3. For combined amount columns with Dr/Cr: split into separate debit and credit values.
4. Parse Indian number formats correctly (e.g., 1,00,000.00 = 100000.00).
5. Return a JSON array of objects with exactly these keys: "date", "narration", "reference", "debit", "credit", "balance"
6. debit, credit, balance should be numbers (float) or null if not applicable.
7. Keep narration as-is from the statement.

Return ONLY a JSON array, no explanation.

Page text:
{page_text[:6000]}"""

    ai_response = await call_claude(parse_prompt)
    parsed = extract_json_from_response(ai_response)

    rows = parsed if isinstance(parsed, list) else parsed.get("transactions", [])
    transactions: List[Transaction] = []
    for row in rows:
        narration = normalize_whitespace(
            " ".join(
                part for part in [
                    str(row.get("narration", "") or ""),
                    str(row.get("reference", "") or ""),
                ] if part
            )
        )
        if "OPENING BALANCE" in narration.upper() or "CLOSING BALANCE" in narration.upper():
            continue
        transactions.append(
            Transaction(
                date=normalize_date(str(row.get("date", ""))),
                narration=narration,
                reference="",
                debit=parse_indian_number(row.get("debit")),
                credit=parse_indian_number(row.get("credit")),
                balance=parse_indian_number(row.get("balance")),
                page_number=page_number,
            )
        )
    return transactions


# ─── Routes ───────────────────────────────────────────────────────────────────

@api_router.get("/")
async def root():
    return {"message": "Bank Statement PDF to Excel Converter API"}


@api_router.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...), password: Optional[str] = Form(None)):
    """Upload PDF, extract text, detect format using AI."""
    try:
        # Validate file type
        if not file.filename or not file.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail="Please upload a PDF file.")

        # Read file into memory
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail=f"File too large. Maximum size is 25MB.")

        if len(content) == 0:
            raise HTTPException(status_code=400, detail="The uploaded file is empty.")

        # Extract text with pdfplumber
        pdf_bytes = io.BytesIO(content)
        try:
            pdf = pdfplumber.open(pdf_bytes, password=password or None)
        except Exception as e:
            error_text = str(e).lower()
            error_repr = repr(e).lower()
            error_type = type(e).__name__.lower()
            is_password_error = (
                "password" in error_text
                or "encrypted" in error_text
                or "passwordincorrect" in error_repr
                or "pdfpasswordincorrect" in error_repr
                or (
                    error_type == "pdfminerexception"
                    and ("passwordincorrect" in error_repr or not error_text.strip())
                )
            )
            if is_password_error:
                if password:
                    raise HTTPException(status_code=400, detail="Incorrect PDF password. Please try again.")
                raise HTTPException(
                    status_code=400,
                    detail="This PDF is password-protected. Enter the PDF password to continue."
                )
            raise HTTPException(status_code=400, detail=f"Could not open PDF: {str(e)}")

        total_pages = len(pdf.pages)
        if total_pages == 0:
            raise HTTPException(status_code=400, detail="The PDF has no pages.")

        # Extract text from all pages
        all_text = []
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text and text.strip():
                all_text.append(f"--- PAGE {i+1} ---\n{text}")

        pdf.close()
        del content  # Free memory

        if not all_text:
            raise HTTPException(status_code=400, detail="No readable text found in the PDF. The document may be scanned/image-based.")

        full_text = "\n".join(all_text)

        # Send first 2 pages to Claude for format detection
        preview_text = "\n".join(all_text[:min(2, len(all_text))])

        detect_prompt = f"""This is text extracted from an Indian bank statement PDF. Analyze it carefully and identify:

(a) bank name if visible
(b) column headers and their exact order as they appear
(c) date format used (e.g., DD/MM/YYYY, DD-MM-YYYY, etc.)
(d) whether debit/credit are separate columns or combined in one column with Dr/Cr suffix
(e) currency symbol used

Return as JSON only with these exact keys:
{{
  "bank_name": "string",
  "columns": ["col1", "col2", ...],
  "date_format": "string",
  "amount_style": "separate" or "combined",
  "currency_symbol": "string"
}}

Bank statement text:
{preview_text[:8000]}"""

        detected = detect_format_locally(preview_text)
        if is_ai_available():
            try:
                ai_response = await call_claude(detect_prompt)
                detected = {**detected, **extract_json_from_response(ai_response)}
            except Exception as e:
                logger.warning(f"AI detection failed, using local detection instead: {e}")

        import base64
        encoded_text = base64.b64encode(full_text.encode('utf-8')).decode('utf-8')

        return {
            "detected_format": {
                "bank_name": detected.get("bank_name", "Unknown Bank"),
                "columns": detected.get("columns", []),
                "date_format": detected.get("date_format", "DD/MM/YYYY"),
                "amount_style": detected.get("amount_style", "separate"),
                "currency_symbol": detected.get("currency_symbol", "₹"),
            },
            "preview_text": preview_text[:3000],
            "total_pages": total_pages,
            "full_text": encoded_text,
            "text_pages_count": len(all_text)
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error processing PDF: {str(e)}")


@api_router.post("/parse-transactions")
async def parse_transactions(data: FormatConfirmation):
    """Parse all transactions using confirmed format."""
    try:
        import base64
        full_text = base64.b64decode(data.full_text).decode('utf-8')
        errors = []

        pages = split_text_into_pages(full_text)
        all_transactions: List[Transaction] = []
        pages_with_transactions = 0
        skipped_pages = 0
        locally_parsed_pages = 0
        ai_fallback_pages = 0

        for page_number, page_text in pages:
            cleaned_page_text = sanitize_page_text(page_text)

            if not page_has_transaction_markers(cleaned_page_text):
                skipped_pages += 1
                continue

            pages_with_transactions += 1
            page_with_marker = f"--- PAGE {page_number} ---\n{cleaned_page_text}"
            page_transactions = parse_transactions_locally(page_with_marker, data.amount_style)
            page_transactions = [txn for txn in page_transactions if txn.page_number == page_number]

            if page_transactions:
                all_transactions.extend(page_transactions)
                locally_parsed_pages += 1
                continue

            if not is_ai_available():
                errors.append(f"Code parser could not extract transactions from page {page_number}.")
                continue

            try:
                ai_transactions = await parse_page_with_ai(cleaned_page_text, page_number, data)
                if ai_transactions:
                    all_transactions.extend(ai_transactions)
                    ai_fallback_pages += 1
                    errors.append(f"Page {page_number} required AI fallback parsing.")
                else:
                    errors.append(f"No transactions found on page {page_number} after AI fallback.")
            except BudgetExceededError as e:
                logger.error(f"Budget exceeded on page {page_number}: {e}")
                errors.append(str(e))
                break
            except Exception as e:
                logger.error(f"Page {page_number} parse error: {e}")
                errors.append(f"Failed to parse page {page_number}: {str(e)}")

        if not all_transactions:
            errors.append("No transaction rows could be parsed from the extracted PDF text.")

        # Balance validation
        for i in range(1, len(all_transactions)):
            prev = all_transactions[i-1]
            curr = all_transactions[i]
            if prev.balance is not None and curr.balance is not None:
                expected = prev.balance
                if curr.debit and curr.debit > 0:
                    expected -= curr.debit
                if curr.credit and curr.credit > 0:
                    expected += curr.credit
                if abs(expected - curr.balance) > BALANCE_TOLERANCE:
                    flipped = False
                    if curr.debit and not curr.credit:
                        flipped_expected = prev.balance + curr.debit
                        if abs(flipped_expected - curr.balance) <= BALANCE_TOLERANCE:
                            all_transactions[i].credit = curr.debit
                            all_transactions[i].debit = None
                            flipped = True
                    elif curr.credit and not curr.debit:
                        flipped_expected = prev.balance - curr.credit
                        if abs(flipped_expected - curr.balance) <= BALANCE_TOLERANCE:
                            all_transactions[i].debit = curr.credit
                            all_transactions[i].credit = None
                            flipped = True

                    if not flipped:
                        all_transactions[i].balance_mismatch = True

        # Compute summary
        total_debit = sum(t.debit or 0 for t in all_transactions)
        total_credit = sum(t.credit or 0 for t in all_transactions)
        opening_balance = all_transactions[0].balance if all_transactions else None
        closing_balance = all_transactions[-1].balance if all_transactions else None

        # Adjust opening balance: first transaction's balance before that txn
        if all_transactions and opening_balance is not None:
            first = all_transactions[0]
            if first.debit and first.debit > 0:
                opening_balance = opening_balance + first.debit
            elif first.credit and first.credit > 0:
                opening_balance = opening_balance - first.credit

        dates = [t.date for t in all_transactions if t.date]
        period_from = dates[0] if dates else ""
        period_to = dates[-1] if dates else ""

        summary = {
            "bank_name": data.bank_name,
            "total_transactions": len(all_transactions),
            "total_debit": round(total_debit, 2),
            "total_credit": round(total_credit, 2),
            "net_flow": round(total_credit - total_debit, 2),
            "opening_balance": round(opening_balance, 2) if opening_balance is not None else None,
            "closing_balance": round(closing_balance, 2) if closing_balance is not None else None,
            "period_from": period_from,
            "period_to": period_to,
            "total_pages": data.total_pages,
            "mismatched_rows": sum(1 for t in all_transactions if t.balance_mismatch),
        }
        confidence = build_parse_confidence(
            transactions=all_transactions,
            total_pages=data.total_pages,
            pages_with_transactions=pages_with_transactions,
            skipped_pages=skipped_pages,
            locally_parsed_pages=locally_parsed_pages,
            ai_fallback_pages=ai_fallback_pages,
            errors=errors,
        )

        return {
            "transactions": [t.model_dump() for t in all_transactions],
            "summary": summary,
            "errors": errors,
            "confidence": confidence.model_dump(),
        }

    except Exception as e:
        logger.error(f"Parse error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error parsing transactions: {str(e)}")


@api_router.post("/download-excel")
async def download_excel(data: ExcelRequest):
    """Generate Excel file from transactions."""
    try:
        wb = openpyxl.Workbook()

        # ─── Sheet 1: Transactions ────────────────────────────────────
        ws = wb.active
        ws.title = "Transactions"

        headers = ["Date", "Particulars", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
        header_fill = PatternFill(start_color="0B2447", end_color="0B2447", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        debit_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
        credit_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        mismatch_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        debit_font = Font(color="DC2626")
        credit_font = Font(color="059669")

        for idx, txn in enumerate(data.transactions, 2):
            ws.cell(row=idx, column=1, value=txn.get("date", "")).border = thin_border
            ws.cell(row=idx, column=2, value=txn.get("narration", "")).border = thin_border
            debit_cell = ws.cell(row=idx, column=3, value=txn.get("debit"))
            debit_cell.number_format = '#,##0.00'
            debit_cell.border = thin_border

            credit_cell = ws.cell(row=idx, column=4, value=txn.get("credit"))
            credit_cell.number_format = '#,##0.00'
            credit_cell.border = thin_border

            balance_cell = ws.cell(row=idx, column=5, value=txn.get("balance"))
            balance_cell.number_format = '#,##0.00'
            balance_cell.border = thin_border

            # Color coding
            is_mismatch = txn.get("balance_mismatch", False)
            is_debit = txn.get("debit") is not None and txn.get("debit", 0) > 0
            is_credit = txn.get("credit") is not None and txn.get("credit", 0) > 0

            if is_mismatch:
                for c in range(1, 6):
                    ws.cell(row=idx, column=c).fill = mismatch_fill
            elif is_debit:
                for c in range(1, 6):
                    ws.cell(row=idx, column=c).fill = debit_fill
                debit_cell.font = debit_font
            elif is_credit:
                for c in range(1, 6):
                    ws.cell(row=idx, column=c).fill = credit_fill
                credit_cell.font = credit_font

        # Auto-fit column widths
        col_widths = [14, 62, 16, 16, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header row and add auto-filters
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:E{len(data.transactions)+1}"

        # ─── Sheet 2: Summary ─────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        summary = data.summary

        summary_header_fill = PatternFill(start_color="0B2447", end_color="0B2447", fill_type="solid")
        summary_header_font = Font(color="FFFFFF", bold=True, size=13)

        ws2.merge_cells('A1:B1')
        cell = ws2.cell(row=1, column=1, value="Bank Statement Summary")
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = Alignment(horizontal="center")
        ws2.cell(row=1, column=2).fill = summary_header_fill

        label_font = Font(bold=True, size=11, color="0B2447")
        value_font = Font(size=11)

        summary_rows = [
            ("Bank Name", summary.get("bank_name", "N/A")),
            ("Statement Period", f"{summary.get('period_from', 'N/A')} to {summary.get('period_to', 'N/A')}"),
            ("Total Transactions", summary.get("total_transactions", 0)),
            ("Total Debits", format_indian_number(summary.get("total_debit", 0))),
            ("Total Credits", format_indian_number(summary.get("total_credit", 0))),
            ("Net Flow", format_indian_number(summary.get("net_flow", 0))),
            ("Opening Balance", format_indian_number(summary.get("opening_balance", 0))),
            ("Closing Balance", format_indian_number(summary.get("closing_balance", 0))),
            ("Balance Mismatches", summary.get("mismatched_rows", 0)),
        ]

        for i, (label, value) in enumerate(summary_rows, 3):
            lbl_cell = ws2.cell(row=i, column=1, value=label)
            lbl_cell.font = label_font
            lbl_cell.border = thin_border
            val_cell = ws2.cell(row=i, column=2, value=value)
            val_cell.font = value_font
            val_cell.border = thin_border

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 35

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"bank_statement_{summary.get('bank_name', 'export').replace(' ', '_')}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        logger.error(f"Excel generation error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
